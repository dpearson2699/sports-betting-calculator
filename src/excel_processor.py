import pandas as pd
from pathlib import Path
from typing import Optional, Tuple, Dict, Any, List, Union

from .betting_framework import user_input_betting_framework
from .commission_manager import commission_manager
from .config.settings import (
    INPUT_DIR, OUTPUT_DIR, DEFAULT_SAMPLE_FILE, 
    DEFAULT_SHEET_NAME, COLUMN_PADDING
)


def get_dynamic_explanation(column_name: str, base_explanation: str) -> str:
    """Generate dynamic explanations that include current commission rate."""
    if column_name in ['Adjusted Price', 'Contract Cost']:
        commission_rate = commission_manager.get_commission_rate()
        # Replace both patterns
        explanation = base_explanation
        if 'contract price + commission' in explanation:
            explanation = explanation.replace('contract price + commission', f'contract price + ${commission_rate:.2f}')
        if '$0.02 commission' in explanation:
            explanation = explanation.replace('$0.02 commission', f'${commission_rate:.2f} commission')
        return explanation
    return base_explanation

# Centralized column configuration
COLUMN_CONFIG: Dict[str, Dict[str, Any]] = {
    # Internal input columns (used for Excel reading and validation)
    'Game': {
        'explanation': 'The game or matchup being analyzed (e.g., "Lakers vs Warriors")',
        'format_type': 'text',
        'is_input': True
    },
    'Model Win Percentage': {
        'explanation': 'Your model\'s predicted win probability (0-100 or 0-1). Input as either 68 (for 68%) or 0.68.',
        'format_type': None,  # Raw input, not formatted in Excel
        'is_input': True,
        'display_name': 'Win %'
    },
    'Model Margin': {
        'explanation': 'Your model\'s predicted margin of victory (optional field for reference only)',
        'format_type': None,
        'is_input': True,
        'display_name': 'Margin'
    },
    'Contract Price': {
        'explanation': 'The price per contract from the sportsbook (e.g., 0.27 for 27 cents, or 27 for 27 cents)',
        'format_type': None,  # Raw input, not formatted in Excel
        'is_input': True,
        'display_name': 'Contract Price (¢)'
    },
    # Display output columns (used in results)
    'Win %': {
        'explanation': 'Your model\'s predicted probability (converted to decimal format for Excel)',
        'format_type': 'percentage',
        'is_input': False
    },
    'Contract Price (¢)': {
        'explanation': 'The input contract price (displayed as entered)',
        'format_type': None,
        'is_input': False
    },
    'Margin': {
        'explanation': 'Your model\'s predicted margin of victory (optional)',
        'format_type': None,
        'is_input': False
    },
    'EV Percentage': {
        'explanation': 'Expected Value percentage - how much profit you expect per dollar wagered. Must be ≥10% (Wharton threshold)',
        'format_type': 'percentage',
        'is_input': False
    },
    'Net Profit': {
        'explanation': 'Total profit you\'ll receive IF this bet wins (what you actually get back minus your stake)',
        'format_type': 'currency',
        'is_input': False
    },
    'Expected Value EV': {
        'explanation': 'Expected profit in dollars accounting for win/loss probability (long-term average over many bets)',
        'format_type': 'currency',
        'is_input': False
    },
    'Decision': {
        'explanation': 'Framework recommendation: BET (meets all criteria) or NO BET (fails Wharton requirements)',
        'format_type': None,
        'is_input': False
    },
    'Final Recommendation': {
        'explanation': 'Final decision after bankroll allocation (may be PARTIAL BET or SKIP if insufficient funds)',
        'format_type': None,
        'is_input': False
    },
    'Bet Amount': {
        'explanation': 'Actual recommended bet amount in dollars, adjusted for whole contracts only (Robinhood constraint)',
        'format_type': 'currency',
        'is_input': False
    },
    'Cumulative Bet Amount': {
        'explanation': 'Actual allocated amount after considering total weekly bankroll constraints',
        'format_type': 'currency',
        'is_input': False
    },
    'Bet Percentage': {
        'explanation': 'Percentage of your weekly bankroll this bet represents (capped at 15% maximum)',
        'format_type': 'percentage',
        'is_input': False
    },
    'Contracts To Buy': {
        'explanation': 'Number of whole contracts to purchase (rounded down from optimal Kelly amount)',
        'format_type': None,
        'is_input': False
    },
    'Adjusted Price': {
        'explanation': 'Total cost per contract including commission (contract price + commission)',
        'format_type': 'currency',
        'is_input': False
    },
    'Target Bet Amount': {
        'explanation': 'Original Kelly/Wharton calculated bet amount before whole contract adjustment',
        'format_type': 'currency',
        'is_input': False
    },
    'Unused Amount': {
        'explanation': 'Money left over due to whole contract constraint (difference between target and actual)',
        'format_type': 'currency',
        'is_input': False
    },
    'Reason': {
        'explanation': 'Explanation for NO BET decisions (e.g., "EV below 10% threshold", "Negative Kelly")',
        'format_type': None,
        'is_input': False
    },
    'Commission Rate': {
        'explanation': 'Commission rate per contract used in calculations (varies by trading platform)',
        'format_type': 'currency',
        'is_input': False
    },
    'Platform': {
        'explanation': 'Trading platform used for commission rate calculation',
        'format_type': None,
        'is_input': False
    }
}

# Quick view column mapping (original -> shortened)
QUICK_VIEW_MAPPING = {
    'Game': 'Game',
    'Win %': 'Win %',
    'Contract Price (¢)': 'Price (¢)',
    'EV Percentage': 'Edge %',
    'Expected Value EV': 'EV $',
    'Adjusted Price': 'Contract Cost',
    'Cumulative Bet Amount': 'Allocated $',
    'Net Profit': 'Win Profit $',
    'Contracts To Buy': 'Contracts',
    'Bet Percentage': 'Stake % Bankroll',
    'Final Recommendation': 'Final',
    'Reason': 'Reason'
}

def get_required_input_columns() -> List[str]:
    """Get list of required input columns from COLUMN_CONFIG."""
    return [col for col, config in COLUMN_CONFIG.items() 
            if config.get('is_input', False) and col != 'Model Margin']  # Model Margin is optional

def apply_excel_formatting(worksheet: Any, df: pd.DataFrame, format_mapping: Optional[Dict[str, Dict[str, Any]]] = None) -> None:
    """Apply formatting to Excel worksheet based on column types."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, PatternFill, Font
    
    # Use COLUMN_CONFIG if no custom mapping provided
    config = format_mapping or COLUMN_CONFIG
    
    # Define commission-related column highlighting
    commission_highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
    commission_columns = ['Commission Rate', 'Platform', 'Adjusted Price', 'Contract Cost']
    
    for col_name, col_config in config.items():
        if col_name in df.columns:
            # get_loc can return int, slice, or array - we only handle int case
            loc_result = df.columns.get_loc(col_name)
            if isinstance(loc_result, int):
                col_idx = loc_result + 1
            else:
                # Handle edge case where column name isn't unique (shouldn't happen in our data)
                col_idx = 1  # Default to first column if complex result
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            
            # Apply number formatting based on type
            format_type = col_config.get('format_type')
            if format_type == 'percentage':
                for row in range(2, len(df) + 2):
                    worksheet[f"{col_letter}{row}"].number_format = '0.00%'
            elif format_type == 'currency':
                for row in range(2, len(df) + 2):
                    worksheet[f"{col_letter}{row}"].number_format = '$0.00'
            elif format_type == 'text':
                for row in range(2, len(df) + 2):
                    worksheet[f"{col_letter}{row}"].number_format = '@'
            
            # Highlight commission-related columns
            if col_name in commission_columns:
                # Highlight header with bold font and background
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = commission_highlight
                header_cell.font = Font(bold=True)
                
                # Highlight data cells with background only
                for row in range(2, len(df) + 2):
                    worksheet[f"{col_letter}{row}"].fill = commission_highlight
            
            # Apply center alignment to ALL column headers
            header_cell = worksheet[f"{col_letter}1"]
            header_cell.alignment = Alignment(horizontal='center')
            
            # Apply right alignment to data cells for general type columns (format_type: None)
            if col_config.get('format_type') is None:
                for row in range(2, len(df) + 2):  # Start from row 2 (skip header)
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(horizontal='right')
            
            # Add explanatory comment to header
            if 'explanation' in col_config:
                header_cell = worksheet[f"{col_letter}1"]
                # Use dynamic explanation that includes current commission rate
                explanation = get_dynamic_explanation(col_name, col_config['explanation'])
                comment = Comment(explanation, "Wharton Betting Framework")
                comment.width = 300
                comment.height = 100
                header_cell.comment = comment

def adjust_column_widths(worksheet: Any, df: Optional[pd.DataFrame] = None, format_mapping: Optional[Dict[str, Dict[str, Any]]] = None) -> None:
    """
    Auto-adjust column widths using Excel-style autofit functionality.
    
    This function uses character-count estimation to adjust column widths.
    
    Args:
        worksheet: Excel worksheet object to adjust
        df: Optional DataFrame containing the data (for content analysis)
        format_mapping: Optional mapping of column names to format types
    """
    
    # Use legacy method for column width adjustment
    _adjust_column_widths_legacy(worksheet)


def _adjust_column_widths_legacy(worksheet: Any) -> None:
    """
    Column width adjustment using character count estimation.
    
    Uses configuration settings for column width rules and padding.
    """
    
    # Simple column width rules
    column_width_rules = {
        'default': {'min': 8, 'max': 25}
    }
    
    # Use configured padding
    padding = COLUMN_PADDING
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        header_name = str(column[0].value) if column[0].value else ""
        
        # Calculate the actual content width needed
        for cell in column:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Get the width rules for this column (use default if not found)
        rules = column_width_rules.get(header_name, column_width_rules['default'])
        
        # Calculate optimal width: content + padding, within min/max bounds
        content_width = max_length + padding
        optimal_width = max(rules['min'], min(content_width, rules['max']))
        
        # Set column width (handle both real and mock worksheets)
        try:
            worksheet.column_dimensions[column_letter].width = optimal_width
        except (KeyError, AttributeError):
            # For mock objects or if column_dimensions doesn't exist
            pass

def list_available_input_files() -> List[str]:
    """List all Excel files in the input directory."""
    excel_files = list(INPUT_DIR.glob("*.xlsx"))
    return [f.name for f in excel_files if not f.name.startswith("~")]  # Exclude temp files

def get_input_file_path(filename: str) -> Path:
    """Get full path for an input file."""
    return INPUT_DIR / filename

def create_sample_excel_in_input_dir() -> Path:
    """Create a sample Excel file in the input directory."""
    sample_data: Dict[str, List[Union[str, int, float]]] = {
        'Game': [
            'Lakers vs Warriors',
            'Cowboys vs Giants', 
            'Yankees vs Red Sox',
            'Chiefs vs Bills',
            'Celtics vs Heat',
            'Dodgers vs Padres'
        ],
        'Model Win Percentage': [68, 72, 55, 75, 63, 58],
        'Model Margin': [3.5, 7.2, 1.8, 10.5, 4.1, 2.3],
        'Contract Price': [45, 0.40, 52, 0.35, 48, 0.51]  # Mixed format: cents and dollars
    }
    
    df = pd.DataFrame(sample_data)
    sample_file_path = INPUT_DIR / DEFAULT_SAMPLE_FILE
    df.to_excel(str(sample_file_path), sheet_name=DEFAULT_SHEET_NAME, index=False)
    print(f"Sample Excel file created: {sample_file_path}")
    return sample_file_path

def process_betting_excel(
    excel_file_path: Union[str, Path], 
    weekly_bankroll: float, 
    sheet_name: str = DEFAULT_SHEET_NAME
) -> Tuple[Optional[pd.DataFrame], Optional[Path]]:
    """
    Process an Excel file with game data through the betting framework.
    
    Expected Excel columns (in this order):
    - Game: Game identifier (e.g., "Team A vs Team B")
    - Model Win Percentage: Your model's win probability (0-100 or 0-1)
    - Model Margin: Optional predicted margin (for reference)
    - Contract Price: Price per unit from sportsbook (dollars or cents)
                     Examples: 0.27 or 27 (both represent 27 cents)
    
    The script will add these columns:
    - Decision: BET or NO BET
    - EV Percentage: Expected value percentage
    - Bet Amount: Actual bet amount (adjusted for whole contracts)
    - Bet Percentage: Percentage of bankroll
    - Net Profit: Profit if bet wins (what you actually win)
    - Expected Value EV: Expected value in dollars (long-term average)
    - Contracts_To_Buy: Number of whole contracts to purchase
    - Target_Bet_Amount: Original Kelly/Wharton recommended amount
    - Unused_Amount: Amount unused due to whole contract constraint
    - Reason: Reason for NO BET decisions
    - Final_Recommendation: After bankroll allocation
    """
    
    try:
        # Read the Excel file
        print(f"Reading Excel file: {excel_file_path}")
        df = pd.read_excel(str(excel_file_path), sheet_name=sheet_name)
        assert isinstance(df, pd.DataFrame), "Expected DataFrame from read_excel"
        print(f"Found {len(df)} games to analyze")
        
        # Validate required columns using centralized config
        required_columns = get_required_input_columns()
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")
        
        # Process each game through the betting framework
        results: List[Dict[str, Any]] = []
        for _, row in df.iterrows():
            game = row['Game']
            win_pct = row['Model Win Percentage']
            contract_price = row['Contract Price']
            
            # Handle margin value - check if column exists and has non-null value
            win_margin = None
            if 'Model Margin' in df.columns:
                margin_val = row.get('Model Margin', None)
                # Only use the value if it's not null/nan
                if pd.notna(margin_val):
                    win_margin = margin_val
            
            print(f"Processing: {game}")
            
            # Run through betting framework
            result = user_input_betting_framework(
                weekly_bankroll=weekly_bankroll,
                model_win_percentage=win_pct,
                contract_price=contract_price,
                model_win_margin=win_margin,
                commission_per_contract=commission_manager.get_commission_rate()
            )
            
            # Calculate Net Profit (what you win if bet hits, accounting for total cost)
            net_profit = 0
            if result['decision'] == 'BET':
                contracts = result.get('contracts_to_buy', 0)
                adjusted_price = result.get('adjusted_price', 0)
                total_cost = contracts * adjusted_price
                payout_if_win = contracts * 1.0  # $1 per contract if win
                net_profit = payout_if_win - total_cost
            
            # Enhance reason with commission impact details for Excel display
            enhanced_reason = result.get('reason', '')
            if result['decision'] == 'NO BET' and enhanced_reason:
                # Add commission impact context to reasons when relevant
                if 'commission_impact' in result and result['commission_impact'] > 0.5:
                    enhanced_reason += f" [Commission impact: -{result['commission_impact']:.1f}% EV]"
                elif 'commission_increase_pct' in result and result['commission_increase_pct'] > 5:
                    enhanced_reason += f" [Commission adds {result['commission_increase_pct']:.0f}% to min bet]"
            
            # Store results using final column names directly
            result_row: Dict[str, Union[str, float, int]] = {
                'Game': game,
                'Win %': win_pct / 100 if win_pct > 1 else win_pct,  # Convert to decimal if > 1
                'Contract Price (¢)': contract_price,
                'Decision': result['decision'],
                'EV Percentage': result['ev_percentage'] / 100,  # Store as decimal for Excel formatting
                'Bet Amount': result['bet_amount'],
                'Bet Percentage': result.get('bet_percentage', 0) / 100,  # Store as decimal for Excel formatting
                'Net Profit': net_profit,
                'Expected Value EV': result.get('expected_profit', 0),
                'Contracts To Buy': result.get('contracts_to_buy', 0),
                'Adjusted Price': result.get('adjusted_price', 0),
                'Target Bet Amount': result.get('target_bet_amount', result['bet_amount']),
                'Unused Amount': result.get('unused_amount', 0),
                'Reason': enhanced_reason,
                'Final Recommendation': '',  # Will be filled by allocation logic
                'Cumulative Bet Amount': 0.0,   # Will be filled by allocation logic
                'Commission Rate': commission_manager.get_commission_rate(),
                'Platform': commission_manager.get_current_platform()
            }
            
            # Only add Margin column if we have margin data
            if win_margin is not None:
                result_row['Margin'] = win_margin
            
            results.append(result_row)
        
        # Create results DataFrame
        results_df = pd.DataFrame(results)
        
        # Sort by EV percentage (highest first) for bet allocation
        results_df = results_df.sort_values('EV Percentage', ascending=False).reset_index(drop=True)
        
        # Apply bankroll allocation logic
        results_df = apply_bankroll_allocation(results_df, weekly_bankroll)
        
        # Reorder columns for improved readability (inputs -> core metrics -> decisions -> sizing diagnostics -> notes)
        # Only include Margin column if input data had margin values
        has_margin_data = 'Model Margin' in df.columns and not df['Model Margin'].isnull().all()
        
        # Logically organized column order for better analysis flow
        preferred_order = [
            # GROUP 1: Game Identification & Input Data
            'Game', 'Win %', 'Contract Price (¢)',
        ]
        if has_margin_data:
            preferred_order.append('Margin')
        
        preferred_order.extend([
            # GROUP 2: Profitability Analysis (core betting metrics)
            'EV Percentage', 'Expected Value EV', 'Net Profit',
            
            # GROUP 3: Bet Sizing Calculations (Kelly → Adjustment → Allocation)
            'Target Bet Amount',        # Original Kelly calculation
            'Bet Amount',              # After whole contract adjustment  
            'Cumulative Bet Amount',   # After bankroll allocation
            'Bet Percentage',          # % of bankroll
            'Unused Amount',           # Money left due to whole contract constraint
            
            # GROUP 4: Contract Implementation Details
            'Contracts To Buy', 'Adjusted Price',
            
            # GROUP 5: Final Decisions & Explanations
            'Decision', 'Final Recommendation', 'Reason'
        ])
        
        # Reorder columns, keeping any extra columns at the end
        existing_cols = [c for c in preferred_order if c in results_df.columns]
        remaining_cols = [c for c in results_df.columns if c not in existing_cols]
        results_df = results_df[existing_cols + remaining_cols]

        # Save results back to Excel in output directory
        input_file = Path(excel_file_path)
        output_file = OUTPUT_DIR / f"{input_file.stem}_RESULTS.xlsx"
        print(f"\nSaving results to: {output_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Create simplified quick-view sheet FIRST using centralized mapping
            quick_cols = [col for col in QUICK_VIEW_MAPPING.keys() if col in results_df.columns]
            quick_df = results_df[quick_cols].copy().rename(columns=QUICK_VIEW_MAPPING)
            quick_df.to_excel(writer, sheet_name='Quick_View', index=False)
            
            # Apply formatting to quick view sheet
            quick_ws = writer.sheets['Quick_View']
            
            # Create format mapping for quick view columns with explanations
            quick_format_mapping: Dict[str, Dict[str, Any]] = {
                'Game': {
                    'format_type': 'text',
                    'explanation': 'The game or matchup being analyzed (e.g., "Lakers vs Warriors")'
                },
                'Win %': {
                    'format_type': 'percentage',
                    'explanation': 'Your model\'s predicted probability that this team/outcome will win'
                },
                'Price (¢)': {
                    'format_type': None,
                    'explanation': 'Contract price from sportsbook (in cents or dollars)'
                },
                'Edge %': {
                    'format_type': 'percentage',
                    'explanation': 'Expected Value percentage - your edge over the market (must be ≥10% for Wharton threshold)'
                },
                'Stake $': {
                    'format_type': 'currency',
                    'explanation': 'Recommended bet amount (adjusted for whole contracts and safety constraints)'
                },
                'Allocated $': {
                    'format_type': 'currency',
                    'explanation': 'Actual amount allocated after considering total bankroll limits'
                },
                'Stake % Bankroll': {
                    'format_type': 'percentage',
                    'explanation': 'Percentage of your weekly bankroll this bet represents (capped at 15% maximum)'
                },
                'Win Profit $': {
                    'format_type': 'currency',
                    'explanation': 'Total profit you\'ll receive IF this bet wins (payout minus your stake)'
                },
                'EV $': {
                    'format_type': 'currency',
                    'explanation': 'Expected profit in dollars accounting for win/loss probability (long-term average)'
                },
                'Contracts': {
                    'format_type': None,
                    'explanation': 'Number of whole contracts to purchase (rounded down from optimal Kelly amount)'
                },
                'Contract Cost': {
                    'format_type': 'currency',
                    'explanation': 'Total cost per contract including $0.02 commission (contract price + commission)'
                },
                'Final': {
                    'format_type': None,
                    'explanation': 'Final recommendation after bankroll allocation (BET, PARTIAL BET, or SKIP)'
                },
                'Reason': {
                    'format_type': None,
                    'explanation': 'Explanation for NO BET decisions (e.g., "EV below 10% threshold")'
                }
            }
            
            apply_excel_formatting(quick_ws, quick_df, quick_format_mapping)
            adjust_column_widths(quick_ws, quick_df, quick_format_mapping)

            # Create detailed results sheet SECOND
            results_df.to_excel(writer, sheet_name='Betting_Results', index=False)
            
            # Apply formatting using helper function
            worksheet = writer.sheets['Betting_Results']
            apply_excel_formatting(worksheet, results_df)
            adjust_column_widths(worksheet, results_df)
        
        # Display summary
        display_summary(results_df, weekly_bankroll)
        
        return results_df, output_file
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None, None

def apply_bankroll_allocation(df: pd.DataFrame, weekly_bankroll: float) -> pd.DataFrame:
    """
    Apply bankroll allocation logic to ensure total bets don't exceed weekly bankroll.
    Games are prioritized by EV percentage (highest first).
    """
    remaining_bankroll = weekly_bankroll
    cumulative_bet = 0
    
    for index, row in df.iterrows():
        if row['Decision'] == 'BET':
            if remaining_bankroll > 0:
                # Check if we can afford this bet
                bet_amount = row['Bet Amount']
                
                if bet_amount <= remaining_bankroll:
                    # Can afford full bet
                    df.loc[index, 'Final Recommendation'] = 'BET'  # type: ignore[assignment]
                    df.loc[index, 'Cumulative Bet Amount'] = float(bet_amount)  # type: ignore[assignment]
                    remaining_bankroll -= bet_amount
                    cumulative_bet += bet_amount
                else:
                    # Can't afford full bet - could do partial or skip
                    if remaining_bankroll >= (weekly_bankroll * 0.01):  # At least 1% of bankroll
                        df.loc[index, 'Final Recommendation'] = f'PARTIAL BET (${remaining_bankroll:.2f})'  # type: ignore[assignment]
                        df.loc[index, 'Cumulative Bet Amount'] = float(remaining_bankroll)  # type: ignore[assignment]
                        cumulative_bet += remaining_bankroll
                        remaining_bankroll = 0
                    else:
                        df.loc[index, 'Final Recommendation'] = 'SKIP - Insufficient Bankroll'  # type: ignore[assignment]
                        df.loc[index, 'Cumulative Bet Amount'] = 0.0  # type: ignore[assignment]
            else:
                # No bankroll remaining - skip all remaining BET decisions
                df.loc[index, 'Final Recommendation'] = 'SKIP - Insufficient Bankroll'  # type: ignore[assignment]
                df.loc[index, 'Cumulative Bet Amount'] = 0.0  # type: ignore[assignment]
        else:
            # Non-BET decisions (e.g., 'NO BET') pass through unchanged
            df.loc[index, 'Final Recommendation'] = row['Decision']  # type: ignore[assignment]
            df.loc[index, 'Cumulative Bet Amount'] = 0.0  # type: ignore[assignment]
    
    return df

def display_summary(df: pd.DataFrame, weekly_bankroll: float) -> None:
    """Display a summary of the betting analysis"""
    print("\n" + "="*60)
    print("BETTING ANALYSIS SUMMARY")
    print("="*60)
    
    total_games = len(df)
    bet_opportunities = len(df[df['Decision'] == 'BET'])
    no_bet_games = len(df[df['Decision'] == 'NO BET'])
    final_bets = len(df[df['Final Recommendation'] == 'BET'])
    
    total_allocated = df['Cumulative Bet Amount'].sum()
    total_expected_profit = df[df['Final Recommendation'] == 'BET']['Expected Value EV'].sum()
    
    print(f"Total Games Analyzed: {total_games}")
    print(f"Initial BET Opportunities: {bet_opportunities}")
    print(f"NO BET Games: {no_bet_games}")
    print(f"Final Recommended Bets: {final_bets}")
    print(f"Weekly Bankroll: ${weekly_bankroll:.2f}")
    print(f"Total Allocated: ${total_allocated:.2f}")
    print(f"Remaining Bankroll: ${weekly_bankroll - total_allocated:.2f}")
    print(f"Total Expected Profit: ${total_expected_profit:.2f}")
    print("-" * 60)
    print("COMMISSION IMPACT ANALYSIS:")
    print(f"Platform: {commission_manager.get_current_platform()}")
    print(f"Commission Rate: ${commission_manager.get_commission_rate():.2f} per contract")
    
    # Calculate total commission costs
    total_contracts = df[df['Final Recommendation'] == 'BET']['Contracts To Buy'].sum()
    total_commission_cost = total_contracts * commission_manager.get_commission_rate()
    
    if total_commission_cost > 0:
        commission_pct_of_bets = (total_commission_cost / total_allocated) * 100 if total_allocated > 0 else 0
        print(f"Total Contracts: {total_contracts}")
        print(f"Total Commission Cost: ${total_commission_cost:.2f}")
        print(f"Commission as % of Total Bets: {commission_pct_of_bets:.1f}%")
        
        # Show commission impact on profitability
        profit_after_commission = total_expected_profit
        profit_before_commission = profit_after_commission + total_commission_cost
        if profit_before_commission > 0:
            commission_impact_pct = (total_commission_cost / profit_before_commission) * 100
            print(f"Commission reduces expected profit by: {commission_impact_pct:.1f}%")
    
    if final_bets > 0:
        print(f"\nTOP 5 RECOMMENDED BETS (by EV%):")
        print("-" * 40)
        top_bets = df[df['Final Recommendation'] == 'BET'].head()
        for _, row in top_bets.iterrows():
            ev_pct = row['EV Percentage'] * 100  # Convert back to percentage for display
            print(f"{row['Game']}: ${row['Cumulative Bet Amount']:.2f} (EV: {ev_pct:.2f}%)")

def create_sample_excel() -> Path:
    """Legacy function - redirects to new create_sample_excel_in_input_dir()"""
    return create_sample_excel_in_input_dir()